import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from api.assessment.constants import ROLES, COLOR_MAPPING
from matplotlib.patheffects import withStroke
import matplotlib.patheffects as path_effects
import matplotlib
matplotlib.use('agg')
from matplotlib import pyplot as plt
import numpy as np
import excel2img
import img2pdf

assets_folder = os.path.join(os.getcwd(), "assessment","assets","level1")

def add_image_to_worksheet(worksheet, folder_path, image_filename, row, column, width, height):
    img = Image(os.path.join(folder_path, image_filename))
    cell = worksheet.cell(row=row, column=column)
    img.width = width
    img.height = height

    worksheet.add_image(img, cell.coordinate)

def Generate_level1_Report(input_label, input_percentages, inclines,virtues,job_info,user_profile,type='career'):
    

    print(inclines)
    # return
    random_hash = str(uuid.uuid4().hex)
    new_folder_path = create_output_folder(random_hash)
    # print(type)
    labels = input_label
    percentages = input_percentages

    create_horizontal_bar_chart(labels, percentages, new_folder_path)
    create_line_chart(virtues, new_folder_path)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page2', inclines=inclines[0])
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page3', inclines=inclines[1],excel=excel)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page6', inclines=inclines[2],excel=excel)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page4',excel=excel,virtues=virtues)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page5',excel=excel,job_info=job_info,user_profile=user_profile)
    excel.save(os.path.join(new_folder_path, "level1report.xlsx"))

    convert_excel_to_pdf(new_folder_path,type)

    return os.path.join(new_folder_path, "output.pdf")


def update_worksheet_cells(worksheet, replacements):
    
    for cell_reference, replacement_value in replacements.items():
        # print(cell_reference)
        
        if isinstance(replacement_value, list):
            param,newValue = replacement_value
            worksheet[cell_reference] = worksheet[cell_reference].value.replace(param,newValue)

        else :
            worksheet[cell_reference] = replacement_value


def create_output_folder(folder_name):
    folder_path = os.path.join(os.getcwd(), "assessment","reports","level1", folder_name)
    os.makedirs(folder_path)
    return folder_path

def create_horizontal_bar_chart(labels, percentages, folder_path):
    for k in range(3):
        fig, ax = plt.subplots(figsize=(60, 20))
        coverted_percentages = [i/72*100 for i in percentages][::-1]
        bar_colors = [(COLOR_MAPPING[label.lower()]) for label in labels]
        bars = ax.barh(labels, coverted_percentages, color=bar_colors[::-1], edgecolor='none',height=0.7)
        ax.set_xlim(0, 100)

        for i,(bar, percentage) in enumerate(zip(bars, percentages)):
            width = bar.get_width() - 5
            ax.text(width, bar.get_y() + bar.get_height() / 2, f'{percentage}', va='center')

            if not i ==2-k:
                bar.set_alpha(0.1)
            else:
                bar.set_alpha(1)

        plt.tight_layout()
        
        plt.savefig(os.path.join(folder_path, f"top3dimmensions-{k+1}.png"))
        plt.clf()
        plt.close()



def create_line_chart(virtues, folder_path):
    variables = [i["virtue"] for i in virtues]
    percentage_values = [round(i["rank"]*100/40) for i in virtues]

    x = range(1, len(variables) + 1)
    plt.figure(figsize=(6, 3))
    plt.plot(x, percentage_values, marker='o', linestyle='-', color='b')

    plt.xlabel("Virtues")
    plt.ylabel("Percentage")
    plt.ylim(0,100)
    plt.xticks(x, variables, rotation=45)  # Set variable names as x-axis labels with 45-degree rotation4

    y_ticks = [i for i in range(0, 101, 10)]
    # y_labels = [f"{y}-{y+10}" for y in range(0, 100, 10)]

    # plt.yticks(y_ticks)

    plt.grid()


    plt.tight_layout()
    plt.savefig(os.path.join(folder_path, "virtueschart.png"), dpi=300,format="png",  transparent=True)
    plt.clf()
    plt.close()

def insert_image_into_excel(folder_path, labels, percentages, worksheet_name, inclines=None,excel=None,virtues=None,job_info=[],user_profile=None):
    if excel:
        workbook = excel
    else:

        workbook = load_workbook(os.path.join(os.getcwd(), "assessment","sample", "level1.xlsx"))
    worksheet = workbook[worksheet_name]

    if worksheet_name == 'Page2':
        add_image_to_worksheet(worksheet,folder_path,"top3dimmensions-1.png",8,7,330,130)
        replacements = {
            "D8": labels[0],
            "D10": labels[1],
            "D11": labels[2],
            "F8": round(percentages[0] * 100 / 72),
            "F10": round(percentages[1] * 100 / 72),
            "F11": round(percentages[2] * 100 / 72)
        }

        update_worksheet_cells(worksheet,replacements)
        update_page2_cells(worksheet, inclines)
    elif worksheet_name == 'Page3':
        add_image_to_worksheet(worksheet,folder_path,"top3dimmensions-2.png",7,7,330,130)

        replacements = {
            "D7": labels[0],
            "D9": labels[1],
            "D11": labels[2],
            "F7": round(percentages[0] * 100 / 72),
            "F9": round(percentages[1] * 100 / 72),
            "F11": round(percentages[2] * 100 / 72)
        }

        update_worksheet_cells(worksheet,replacements)
        update_page3_cells(worksheet, inclines)
    elif worksheet_name == 'Page6':
        add_image_to_worksheet(worksheet,folder_path,"top3dimmensions-3.png",7,7,330,130)

        replacements = {
            "D7": labels[0],
            "D9": labels[1],
            "D11": labels[2],
            "F7": round(percentages[0] * 100 / 72),
            "F9": round(percentages[1] * 100 / 72),
            "F11": round(percentages[2] * 100 / 72)
        }

        update_worksheet_cells(worksheet,replacements)

        update_page3_cells(worksheet, inclines)
    elif worksheet_name == 'Page4':
        add_image_to_worksheet(worksheet,folder_path,"virtueschart.png",7,7,350,200)
        update_page4_cells(worksheet,virtues,folder_path)
    elif worksheet_name == 'Page5':
        update_page5_cells(worksheet,folder_path,labels,job_info,user_profile)
        

    
    return workbook

    

def update_page2_cells(worksheet, inclines):
    replacements = {
            "D15": ['input_dimension',inclines['feature']],
            "D19":inclines['purpose_statement'],
            "D24":"You" + inclines['thrive_environment'],
            "D42":inclines['career_inclination_statement'],
            "D46":inclines['quote'],
    }
    
    

    inclinations = inclines['inclinations'].split('\n')
    cont = 30
    for i in range(len(inclinations)):
        replacements[f"D{cont}"] = inclinations[i] 
        cont = cont + 2   
    
    
    # careers = inclines['careers'].split(",")[:3]
    # for i in range(46, 49):
    #     replacements[f"E{i}"] = careers[i - 46].strip() 
    update_worksheet_cells(worksheet,replacements)
    add_image_to_worksheet(worksheet,assets_folder,f"{inclines['feature']}.png",10,7,250,325)

def update_page3_cells(worksheet, inclines):
    # pass
    replacements = {
            "B8": ['input_dimension',inclines['feature']],
            "B13":inclines['purpose_statement'],
            "B16":"You" + inclines['thrive_environment'],
            "B34":inclines['career_inclination_statement'],
            "B40":inclines['quote']
    }



    # worksheet["B10"] = worksheet["B10"].value.replace("input_dimension", inclines['feature'])

    inclinations = inclines['inclinations'].split('\n')
    cont = 21
    for i in range(len(inclinations)):
        replacements[f"B{cont}"] = inclinations[i]
        cont = cont + 2
    
    # careers = inclines['careers'].split(",")[:5]
    # for i in range(37, 40):
    #     replacements[f"C{i}"] = careers[i - 37]
    

    update_worksheet_cells(worksheet,replacements)
    add_image_to_worksheet(worksheet,assets_folder,f"{inclines['feature']}.png",14,7,250,325)
    


def update_page6_cells(worksheet, inclines):

    replacements = {
        "B7":["input_dimension",inclines['feature']],
        "B13":inclines['purpose_statement'],
        "B16":"You"+ inclines['thrive_environment'],
        "B34":inclines['career_inclination_statement'],
        "B40":inclines['quote']


    }
    inclinations = inclines['inclinations'].split('\n')
    cont = 21
    for i in range(len(inclinations)):
        replacements[f"B{cont}"] = inclinations[i]
        cont = cont + 2 

    # careers = inclines['careers'].split(",")[:5]
    # for i in range(37, 40):
    #     replacements[f"C{i}"] = careers[i - 37]
    
    update_worksheet_cells(worksheet,replacements)


def update_page4_cells(worksheet,virtues,folder_path):
    cell_positions = [("B15", "B14"), ("E19", "E18"), ("B22", "B21")]

    for i, (value_cell, virtue_cell) in enumerate(cell_positions):
        worksheet[value_cell] = virtues[i]['text']
        worksheet[virtue_cell] = virtues[i]['virtue']


    

    for i in range(27,30):
        worksheet[f"B{i}"] = f'{i-26}. {virtues[i-24]["virtue"]}'
    
    for i in range(37,40):
        worksheet[f"B{i}"] = f'{i-36}. {virtues[i-31]["virtue"]}'

    cell_coordinates = [(14, 6), (18, 2), (21, 6)]
    for i in range(3):
        data = [virtues[i]['rank'] * 2]

        fig, ax = plt.subplots(figsize=(7, 4))

        ax.barh(1, data, color='#dddbd1', height=0.5, align='center')

        ax.set_xlabel('X-axis Label')
        ax.set_ylabel('Y-axis Label')
        ax.set_xlim(0, 100)
        
        ax.set_yticks([])

        plt.savefig(os.path.join(folder_path, f"virtue{i + 1}.png"), dpi=300, transparent=True,  bbox_inches='tight'    )
        row, col = cell_coordinates[i]
        add_image_to_worksheet(worksheet,folder_path,f"virtue{i+1}.png",row,col,350,95)



def update_page5_cells(worksheet,folder_path,labels,job_info,user_profile):
   
    # worksheet["B7"] = worksheet["B7"].value.replace("input_dimension", inclines['feature'])
    # worksheet["B10"] = worksheet["B10"].value.replace("input_dimension", inclines['feature'])
   

    # worksheet["B15"] = job_info[0]['career_cluster']
    # worksheet["C15"] = job_info[0]['job_name']
    # worksheet["E15"] = job_info[0]['lwdimension_field1']
    # worksheet["E16"] = job_info[0]['lwdimension_field2']
    # worksheet["E17"] = job_info[0]['lwdimension_field3']


    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    recipe = list(COLOR_MAPPING.keys())
    data = [1]*len(recipe)

    colors = [COLOR_MAPPING[recipe_name] for recipe_name in recipe]

    # Create an "explode" list to specify which segments to explode
    explode = [0.05, 0.05, 0.05, 0.05, 0.05, 0.05,0.05, 0.05, 0.05]

    wedges, texts = ax.pie(data, colors=colors, wedgeprops=dict(width=0.15), startangle=-40, explode=explode)

    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-"), bbox=bbox_props, zorder=0, va="center")

    # Add annotations only for "angel" and "principal"

    lw_dimensions = job_info[0]
    dimension_fields = ['lwdimension_field1', 'lwdimension_field2', 'lwdimension_field3']

    annotation_indices = [recipe.index(lw_dimensions[dim].lower()) for dim in dimension_fields]

    

    for i, p in enumerate(wedges):
        if i in annotation_indices:
            ang = (p.theta2 - p.theta1) / 2. + p.theta1
            y = np.sin(np.deg2rad(ang))
            x = np.cos(np.deg2rad(ang))
            horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
            connectionstyle = f"angle,angleA=0,angleB={ang}"
            kw["arrowprops"].update({"connectionstyle": connectionstyle})
            ax.annotate(recipe[i], xy=(x, y), xytext=(1.75 * np.sign(x), 1.4 * y),
                        horizontalalignment=horizontalalignment, **kw)
            
            worksheet[f"B{23+annotation_indices.index(i)}"] = recipe[i].capitalize()
            worksheet[f"C{23+annotation_indices.index(i)}"] = ROLES[recipe[i].lower()]

        # else:
        #     ang = (p.theta2 - p.theta1) / 2. + p.theta1
        #     y = np.sin(np.deg2rad(ang))
        #     x = np.cos(np.deg2rad(ang))
        #     horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        #     ax.text(x, y, recipe[i], horizontalalignment=horizontalalignment, fontsize=10, color='black')

            # p.set_alpha(1)  # Hide wedges that are not annotated

    # Add 'A' to the center of the pie chart with path effects
    center_text = plt.gca().text(0.0, 0.0, 'Limeneal', color='white', ha='center', va='center', size=10,fontfamily='sans-serif')
    stroke = withStroke(linewidth=3, foreground='black')
    center_text.set_path_effects([stroke, path_effects.Normal()])

    center_circle = plt.Circle((0, 0), 0.35, color='#f5cff7')
    ax.add_artist(center_circle)
    plt.tight_layout()

    plt.savefig(os.path.join(folder_path, "job1dimmensions.png"), dpi=300,transparent=True)
    plt.clf()
    plt.close()

     # worksheet["B12"] = inclines['purpose_statement']4
    
    add_image_to_worksheet(worksheet,folder_path,"job1dimmensions.png",12,2,500,230)

    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
    explode = [0.05, 0.05, 0.05, 0.05, 0.05, 0.05,0.05, 0.05, 0.05]

    wedges, texts = ax.pie(data, colors=colors, wedgeprops=dict(width=0.15), startangle=-40, explode=explode)

    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-"), bbox=bbox_props, zorder=0, va="center")


    lw_dimensions = job_info[0]
    

    annotation_indices = [recipe.index(i.lower()) for i in labels]

    for i, p in enumerate(wedges):
        if i in annotation_indices:
            ang = (p.theta2 - p.theta1) / 2. + p.theta1
            y = np.sin(np.deg2rad(ang))
            x = np.cos(np.deg2rad(ang))
            horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
            connectionstyle = f"angle,angleA=0,angleB={ang}"
            kw["arrowprops"].update({"connectionstyle": connectionstyle})
            ax.annotate(recipe[i], xy=(x, y), xytext=(1.75 * np.sign(x), 1.4 * y),
                        horizontalalignment=horizontalalignment, **kw)

           
    center_text = plt.gca().text(0.0, 0.0, 'Limeneal', color='white', ha='center', va='center', size=10,fontfamily='sans-serif')
    stroke = withStroke(linewidth=3, foreground='black')
    center_text.set_path_effects([stroke, path_effects.Normal()])

    center_circle = plt.Circle((0, 0), 0.35, color='#f5cff7')
    ax.add_artist(center_circle)
    # plt.title(f"Sumith's Inclination'")
    plt.tight_layout()

    plt.savefig(os.path.join(folder_path, "usersdimmensions.png"), dpi=300,transparent=True)
    plt.clf()
    plt.close()

    add_image_to_worksheet(worksheet,folder_path,"usersdimmensions.png",12,6,500,230)
    
    replacements = {
        "C11": f"{job_info[0]['job_name']}'s Inclinations",
        "G11": f"{user_profile.name}'s Inclinations",
        "B21": ['job_name',job_info[0]['job_name']],
        "F21": ['user_name',user_profile.name]
    }

    update_worksheet_cells(worksheet,replacements)
    for i in range(23,26):
        worksheet[f"F{i}"] = labels[i-23].capitalize()
        worksheet[f"G{i}"] =   ROLES[labels[i-23].lower()]



def convert_excel_to_pdf(folder_path, type, excel_filename="level1report.xlsx", num_pages=4, page_prefix="Page"):
    
    image_paths = []
    if type == "Career":
        pages_to_include = [1, 2, 3,6, 4, 5]
    elif type == "Leadership":
        pages_to_include = [1, 2, 4]
    else:
        pages_to_include = list(range(1, num_pages + 1))

    for page_num in pages_to_include:
        page_name = f"{page_prefix}{page_num}"
        image_name = f"{page_name}.png"
        excel2img.export_img(
            os.path.join(folder_path, excel_filename),
            os.path.join(folder_path, image_name),
            page_name,
            "A1:H53"
        )
        image_paths.append(os.path.join(folder_path, image_name))

   
    pdf_data = img2pdf.convert(image_paths)

    pdf_output_path = os.path.join(folder_path, "output.pdf")
    with open(pdf_output_path, "wb") as file:
        file.write(pdf_data)

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if filename != "output.pdf" and os.path.isfile(file_path):
            os.remove(file_path)
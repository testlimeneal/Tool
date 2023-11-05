import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from api.assessment.constants import ROLES, COLOR_MAPPING
from matplotlib.patheffects import withStroke
import matplotlib.patheffects as path_effects
import matplotlib
matplotlib.use('agg')
from matplotlib import pyplot as plt
import numpy as np
import excel2img
import img2pdf


activities_path = os.path.join(os.getcwd(), "assessment","assets","level2", 'dimmensions')
assets_folder = os.path.join(os.getcwd(), "assessment","assets","level1")

emotions_path = os.path.join(os.getcwd(), "assessment","assets","level2", 'emotions')
virtues_path = os.path.join(os.getcwd(), "assessment","assets","level2", 'virtues')


def add_image_to_worksheet(worksheet, folder_path, image_filename, row, column, width, height):
    img = Image(os.path.join(folder_path, image_filename))
    cell = worksheet.cell(row=row, column=column)
    img.width = width
    img.height = height
    worksheet.add_image(img, cell.coordinate)

def create_output_folder(folder_name):
    folder_path = os.path.join(os.getcwd(), "assessment","reports","level2", folder_name)
    os.makedirs(folder_path)
    return folder_path

def update_worksheet_cells(worksheet, replacements):
    for cell_reference, replacement_value in replacements.items():
        
        if isinstance(replacement_value, list):
            param,newValue = replacement_value
            worksheet[cell_reference] = worksheet[cell_reference].value.replace(param,newValue)

        else :
            worksheet[cell_reference] = replacement_value



def Generate_level2_Report(res,nlp_data,bucket_mapping):
    
    random_hash = str(uuid.uuid4().hex)
    new_folder_path = create_output_folder(random_hash)

    # labels = input_label
    # percentages = input_percentages

    # create_horizontal_bar_chart(labels, percentages, new_folder_path)
    # create_line_chart(virtues, new_folder_path)
    # print(res)
    # return
    

    excel = insert_image_into_excel(worksheet_name='Page2',data=res)
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page3',data=[res[2],nlp_data])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page4',data=[res[0][0],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page5',data=[res[0][1],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page6',data=[res[0][2],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page7',data=[res[1][0],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page8',data=[res[1][1],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page9',data=[res[1][2],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page10',data=[res[2][0],bucket_mapping,'pain'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page11',data=[res[2][1],bucket_mapping,'pain'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page12',data=[res[2][2],bucket_mapping,'pain'])
    # print(res)
    # return
    # excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page3', inclines=inclines[1],excel=excel)
    # excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page6', inclines=inclines[2],excel=excel)
    # excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page4',excel=excel,virtues=virtues)
    # excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page5',excel=excel,job_info=job_info)
    excel.save(os.path.join(new_folder_path, "level2report.xlsx"))

    convert_excel_to_pdf(new_folder_path,"type")

    return os.path.join(new_folder_path, "output.pdf")







def create_horizontal_bar_chart(labels, percentages, folder_path):
    fig, ax = plt.subplots(figsize=(60, 20))
    coverted_percentages = [i/80*100 for i in percentages][::-1]
    bar_colors = [(COLOR_MAPPING[label.lower()]) for label in labels]
    bars = ax.barh(labels, coverted_percentages, color=bar_colors[::-1], edgecolor='none',height=0.7)
    ax.set_xlim(0, 100)

    for bar, percentage in zip(bars, percentages):
        width = bar.get_width() - 5
        ax.text(width, bar.get_y() + bar.get_height() / 2, f'{percentage}', va='center')

    plt.tight_layout()
    
    plt.savefig(os.path.join(folder_path, "top3dimmensions.png"), dpi=300)
    plt.clf()
    plt.close()



def create_line_chart(virtues, folder_path):
    variables = [i["virtue"] for i in virtues]
    percentage_values = [round(i["rank"]*100/40) for i in virtues]

    x = range(1, len(variables) + 1)
    plt.figure(figsize=(6, 3))
    plt.plot(x, percentage_values, marker='o', linestyle='-', color='b')

    plt.xlabel("Virtues")
    plt.ylabel("Percentage")
    plt.ylim(0,100)
    plt.xticks(x, variables, rotation=45)  # Set variable names as x-axis labels with 45-degree rotation
    plt.grid()


    plt.tight_layout()
    plt.savefig(os.path.join(folder_path, "virtueschart.png"), dpi=300,format="png",  transparent=True)
    plt.clf()
    plt.close()

def insert_image_into_excel(worksheet_name, data=None,excel=None):
    if excel:
        workbook = excel
    else:

        workbook = load_workbook(os.path.join(os.getcwd(), "assessment","sample", "level2.xlsx"))
    worksheet = workbook[worksheet_name]

    if worksheet_name == 'Page2':
        replacements = {}


        sr_no = 22
        cell_row = ['B27', 'C27', 'G27','B42', 'C42', 'G42']
        for sublist in data:
            for item in sublist:
                # print(item)
                replacements[f"L{sr_no}"] = item['name']
                replacements[f"M{sr_no}"] = round(item['value'])
                if sr_no-22 < 6:
                    replacements[cell_row[sr_no-22]] = ROLES[item['name'].lower()]
                    # print(ROLES[item['name'].lower()])
                sr_no = sr_no + 1
        
        update_worksheet_cells(worksheet,replacements)
        # add_image_to_worksheet(worksheet,folder_path,"top3dimmensions.png",8,7,330,130)  
        # update_page2_cells(worksheet, inclines)
    elif worksheet_name == 'Page3':
        
        replacements = {}

        res_data,nlp_data = data

        cell_row = ['B18','C18','G18']
        for i in range(3):
            replacements[cell_row[i]] = ROLES[res_data[i]['name'].lower()]
        
        cell_row = [('B27','B29','B34','B35'),('B40','B42','B47','B48')]
        for i,j in enumerate(nlp_data):
            replacements[cell_row[i][0]] = j['name'] 
            replacements[cell_row[i][1]] = j['statement'] 
            replacements[cell_row[i][2]] = ' '.join(j['learnings'].split()[:2]) 
            replacements[cell_row[i][3]] = ' '.join(j['learnings'].split()[2:]) 
        
        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name == 'Page4':
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]

        b35_cell = worksheet['B35'].value.replace('input_emotion',feature_data['emotion']).replace('input_colour',feature_data['colour'])
        c39_cell = worksheet['C39'].value.replace('input_virtue',feature_data['virtue']).replace('input_dimmension',bucket['name'])

        replacements = {
            'B12': bucket['name'],
            'G10' : feature_data[f'{type}_motivation'],
            'B30' : feature_data['motivation'],
            'B33' : ['input_dimmension',bucket['name']],
            'B34' : ['input_dimmension',bucket['name']],
            'B35' : b35_cell,
            'C39' : c39_cell,
            'C40': feature_data[f'{type}_virtue']
        }

        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{16+i}"] = j

        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"G{16+i}"] = j

        add_image_to_worksheet(worksheet,assets_folder,f"{bucket['name']}.png",7,4,250,325)
        add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",23,6,120,120)
        add_image_to_worksheet(worksheet,emotions_path,f"{bucket['name']}.png",34,7,120,120)
        add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",38,2,120,120)

        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name in ['Page5','Page6']:
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]
        b33_cell = worksheet['B33'].value.replace('input_emotion',feature_data['emotion']).replace('input_colour',feature_data['colour'])
        c36_cell = worksheet['C36'].value.replace('input_virtue',feature_data['virtue']).replace('input_dimmension',bucket['name'])
        replacements = {
            'B8': bucket['name'],
            'G7' : feature_data[f'{type}_motivation'],
            'B28' : feature_data['motivation'],
            'B31' : ['input_dimmension',bucket['name']],
            'B32' : ['input_dimmension',bucket['name']],
            'B33' : b33_cell,
            # 'C39' : ['input_dimmension',bucket['name']],
            'C37': feature_data[f'{type}_virtue'],
            'C36' : c36_cell
        }
        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{13+i}"] = j
        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"G{13+i}"] = j

        add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",20,6,120,120)
        add_image_to_worksheet(worksheet,emotions_path,f"{bucket['name']}.png",30,7,120,120)
        add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",35,2,120,120)
        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name in ['Page7','Page8','Page9','Page10','Page11','Page12']:
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]

        c31_cell = worksheet['C31'].value.replace('input_virtue',feature_data['virtue']).replace('input_dimmension',bucket['name'])

        replacements = {
            'B12': bucket['name'],
            'G10' : feature_data[f'{type}_motivation'],
            # 'B30' : feature_data['motivation'],
            'C31' : c31_cell,
            # 'B35' : b35_cell,
            # 'C39' : c39_cell,
            'C32': feature_data[f'{type}_virtue']
        }

        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{18+i}"] = j

        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"G{18+i}"] = j

        add_image_to_worksheet(worksheet,assets_folder,f"{bucket['name']}.png",7,4,250,325)
        add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",25,6,120,120)
        add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",30,2,120,120)

        update_worksheet_cells(worksheet,replacements)

    #     update_page3_cells(worksheet, inclines)
    # elif worksheet_name == 'Page6':
    #     update_page3_cells(worksheet, inclines)
    # elif worksheet_name == 'Page4':
    #     add_image_to_worksheet(worksheet,folder_path,"virtueschart.png",7,7,350,200)
    #     update_page4_cells(worksheet,virtues,folder_path)
    # elif worksheet_name == 'Page5':
    #     update_page5_cells(worksheet,folder_path,labels,job_info)

    
    return workbook

def convert_excel_to_pdf(folder_path, type, excel_filename="level2report.xlsx", num_pages=14, page_prefix="Page"):
    
    image_paths = []
    # if type == "Career":
    #     pages_to_include = [1, 2, 3,6]
    # elif type == "Leadership":
    #     pages_to_include = [1, 2, 4]
    # else:
    pages_to_include = list(range(1, num_pages + 1))

    for page_num in pages_to_include:
        page_name = f"{page_prefix}{page_num}"
        image_name = f"{page_name}.png"
        excel2img.export_img(
            os.path.join(folder_path, excel_filename),
            os.path.join(folder_path, image_name),
            page_name,
            "A1:H53"
        )
        image_paths.append(os.path.join(folder_path, image_name))

   
    pdf_data = img2pdf.convert(image_paths)

    pdf_output_path = os.path.join(folder_path, "output.pdf")
    with open(pdf_output_path, "wb") as file:
        file.write(pdf_data)

    # for filename in os.listdir(folder_path):
    #     file_path = os.path.join(folder_path, filename)
    #     if filename != "output.pdf" and os.path.isfile(file_path):
    #         os.remove(file_path)